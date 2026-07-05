"""File parser service — extract text from PDF, Word, Excel documents."""

import io
import logging

logger = logging.getLogger(__name__)

MAX_TEXT_LENGTH = 100_000  #字符上限
MAX_PDF_PAGES = 500
MAX_XLSX_ROWS = 10_000


def parse_file(file_bytes: bytes, filename: str) -> dict:
    """Unified entry: detect file type and dispatch to the right parser."""
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext == "pdf":
        return parse_pdf(file_bytes, filename)
    elif ext in ("docx", "doc"):
        return parse_docx(file_bytes, filename)
    elif ext in ("xlsx", "xls"):
        return parse_xlsx(file_bytes, filename)
    else:
        return {"error": f"不支持的文件格式: .{ext}"}


def _truncate(text: str) -> str:
    if len(text) > MAX_TEXT_LENGTH:
        return text[:MAX_TEXT_LENGTH] + f"\n\n…（超出 {MAX_TEXT_LENGTH} 字符限制，已截断）"
    return text


# ── PDF ──────────────────────────────────────────────────────────────

def parse_pdf(file_bytes: bytes, filename: str = "") -> dict:
    """Extract text from PDF using PyMuPDF."""
    try:
        import fitz
    except ImportError:
        return {"error": "PDF 解析不可用（未安装 PyMuPDF）"}

    try:
        doc = fitz.open(stream=file_bytes, filetype="pdf")
    except Exception as e:
        return {"error": f"无法打开 PDF 文件: {e}"}

    total_pages = doc.page_count
    if total_pages > MAX_PDF_PAGES:
        doc.close()
        return {"error": f"PDF 页数过多（{total_pages} 页），限制为 {MAX_PDF_PAGES} 页"}

    pages_text = []
    for page_num in range(total_pages):
        page = doc[page_num]
        text = page.get_text("text").strip()
        if text:
            pages_text.append(f"--- 第 {page_num + 1} 页 ---\n{text}")

    doc.close()

    full_text = "\n\n".join(pages_text)
    if not full_text:
        return {
            "filename": filename,
            "text": "",
            "pages": total_pages,
            "size": 0,
            "warning": "此 PDF 未提取到文字内容（可能为扫描件），建议截图后使用图片识别功能",
        }

    truncated = _truncate(full_text)
    return {
        "filename": filename,
        "text": truncated,
        "pages": total_pages,
        "size": len(full_text),
    }


# ── Word ─────────────────────────────────────────────────────────────

def parse_docx(file_bytes: bytes, filename: str = "") -> dict:
    """Extract paragraphs from .docx file."""
    try:
        from docx import Document
    except ImportError:
        return {"error": "Word 解析不可用（未安装 python-docx）"}

    try:
        doc = Document(io.BytesIO(file_bytes))
    except Exception as e:
        return {"error": f"无法打开 Word 文件: {e}"}

    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    full_text = "\n\n".join(paragraphs)

    if not full_text:
        return {
            "filename": filename,
            "text": "",
            "paragraphs": 0,
            "size": 0,
            "warning": "Word 文件中未提取到文字内容",
        }

    truncated = _truncate(full_text)
    return {
        "filename": filename,
        "text": truncated,
        "paragraphs": len(paragraphs),
        "size": len(full_text),
    }


# ── Excel ────────────────────────────────────────────────────────────

def parse_xlsx(file_bytes: bytes, filename: str = "") -> dict:
    """Extract cell contents from .xlsx file, one sheet at a time."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"error": "Excel 解析不可用（未安装 openpyxl）"}

    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        return {"error": f"无法打开 Excel 文件: {e}"}

    sheet_blocks = []
    total_rows = 0

    for sheet_name in wb.sheetnames:
        if total_rows >= MAX_XLSX_ROWS:
            break
        ws = wb[sheet_name]
        rows_text = []
        row_count = 0
        for row in ws.iter_rows(values_only=True):
            if total_rows >= MAX_XLSX_ROWS:
                rows_text.append(f"\n…（超过 {MAX_XLSX_ROWS} 行限制，已截断）")
                break
            cells = [str(c) for c in row if c is not None]
            if cells:
                rows_text.append(" | ".join(cells))
                row_count += 1
                total_rows += 1

        if rows_text:
            sheet_blocks.append(f"[Sheet: {sheet_name}]\n" + "\n".join(rows_text))

    all_sheets = list(wb.sheetnames)
    wb.close()

    full_text = "\n\n".join(sheet_blocks)
    if not full_text:
        return {
            "filename": filename,
            "text": "",
            "sheets": len(all_sheets),
            "rows": 0,
            "size": 0,
            "warning": "Excel 文件中未提取到数据",
        }

    truncated = _truncate(full_text)
    return {
        "filename": filename,
        "text": truncated,
        "sheets": len(all_sheets),
        "rows": total_rows,
        "size": len(full_text),
    }
